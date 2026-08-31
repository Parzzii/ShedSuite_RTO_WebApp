from __future__ import annotations

import asyncio
import base64
import ctypes
import json
import os
import re
import tempfile
import threading
from pathlib import Path

import fitz
from openpyxl import load_workbook
from playwright.async_api import async_playwright, expect

from pdf_tools import bytes_to_pdf, safe_filename

BASE_URL = 'https://app.shedsuite.com/rto/order/'
HOME_URL = 'https://app.shedsuite.com/'

# V7.22: in-app ShedSuite account vault + persistent company -> login mapping.
# New companies are automatically probed against saved accounts and the correct
# tenant is remembered. The legacy Logininfo workbook remains a fallback only.
_LOGIN_MAPPING_LOCK = threading.RLock()
_ACCOUNT_STORE_LOCK = threading.RLock()


def _account_store_path() -> Path:
    """Cross-version local account store managed by the web app."""
    override = str(os.getenv('SHEDSUITE_ACCOUNTS_FILE', '') or '').strip()
    if override:
        return Path(override).expanduser()
    if os.name == 'nt':
        root = Path(os.getenv('LOCALAPPDATA') or Path.home())
        return root / 'ShedSuiteRTO' / 'shedsuite_accounts.json'
    return Path.home() / '.shedsuite_rto' / 'shedsuite_accounts.json'


def _dpapi_protect(raw: bytes) -> bytes:
    """Protect bytes with Windows DPAPI for the current Windows user."""
    if os.name != 'nt':
        return raw

    class DATA_BLOB(ctypes.Structure):
        _fields_ = [('cbData', ctypes.c_ulong), ('pbData', ctypes.POINTER(ctypes.c_ubyte))]

    crypt32 = ctypes.windll.crypt32
    kernel32 = ctypes.windll.kernel32
    crypt32.CryptProtectData.argtypes = [
        ctypes.POINTER(DATA_BLOB), ctypes.c_wchar_p, ctypes.POINTER(DATA_BLOB),
        ctypes.c_void_p, ctypes.c_void_p, ctypes.c_ulong, ctypes.POINTER(DATA_BLOB),
    ]
    crypt32.CryptProtectData.restype = ctypes.c_bool
    kernel32.LocalFree.argtypes = [ctypes.c_void_p]
    kernel32.LocalFree.restype = ctypes.c_void_p
    in_buf = ctypes.create_string_buffer(raw)
    in_blob = DATA_BLOB(len(raw), ctypes.cast(in_buf, ctypes.POINTER(ctypes.c_ubyte)))
    out_blob = DATA_BLOB()
    if not crypt32.CryptProtectData(ctypes.byref(in_blob), 'ShedSuiteRTO', None, None, None, 0, ctypes.byref(out_blob)):
        raise ctypes.WinError()
    try:
        return ctypes.string_at(out_blob.pbData, out_blob.cbData)
    finally:
        kernel32.LocalFree(out_blob.pbData)


def _dpapi_unprotect(raw: bytes) -> bytes:
    if os.name != 'nt':
        return raw

    class DATA_BLOB(ctypes.Structure):
        _fields_ = [('cbData', ctypes.c_ulong), ('pbData', ctypes.POINTER(ctypes.c_ubyte))]

    crypt32 = ctypes.windll.crypt32
    kernel32 = ctypes.windll.kernel32
    crypt32.CryptUnprotectData.argtypes = [
        ctypes.POINTER(DATA_BLOB), ctypes.POINTER(ctypes.c_wchar_p), ctypes.POINTER(DATA_BLOB),
        ctypes.c_void_p, ctypes.c_void_p, ctypes.c_ulong, ctypes.POINTER(DATA_BLOB),
    ]
    crypt32.CryptUnprotectData.restype = ctypes.c_bool
    kernel32.LocalFree.argtypes = [ctypes.c_void_p]
    kernel32.LocalFree.restype = ctypes.c_void_p
    in_buf = ctypes.create_string_buffer(raw)
    in_blob = DATA_BLOB(len(raw), ctypes.cast(in_buf, ctypes.POINTER(ctypes.c_ubyte)))
    out_blob = DATA_BLOB()
    if not crypt32.CryptUnprotectData(ctypes.byref(in_blob), None, None, None, None, 0, ctypes.byref(out_blob)):
        raise ctypes.WinError()
    try:
        return ctypes.string_at(out_blob.pbData, out_blob.cbData)
    finally:
        kernel32.LocalFree(out_blob.pbData)


def _encode_password(password: str) -> str:
    raw = str(password or '').encode('utf-8')
    protected = _dpapi_protect(raw)
    prefix = 'dpapi:' if os.name == 'nt' else 'local:'
    return prefix + base64.b64encode(protected).decode('ascii')


def _decode_password(value: str) -> str:
    value = str(value or '')
    if not value:
        return ''
    if ':' not in value:
        return ''
    prefix, payload = value.split(':', 1)
    try:
        raw = base64.b64decode(payload)
        if prefix == 'dpapi':
            raw = _dpapi_unprotect(raw)
        return raw.decode('utf-8')
    except Exception:
        return ''


def _load_account_store() -> dict:
    path = _account_store_path()
    with _ACCOUNT_STORE_LOCK:
        if not path.exists():
            return {'version': 1, 'accounts': []}
        try:
            data = json.loads(path.read_text(encoding='utf-8'))
        except Exception:
            return {'version': 1, 'accounts': []}
        if not isinstance(data, dict):
            return {'version': 1, 'accounts': []}
        if not isinstance(data.get('accounts'), list):
            data['accounts'] = []
        return data


def _save_account_store(data: dict) -> None:
    path = _account_store_path()
    with _ACCOUNT_STORE_LOCK:
        path.parent.mkdir(parents=True, exist_ok=True)
        temp = path.with_suffix('.tmp')
        temp.write_text(json.dumps(data, indent=2), encoding='utf-8')
        temp.replace(path)


def list_shedsuite_accounts(base: Path | None = None, include_legacy: bool = True) -> list[dict]:
    """Return safe account metadata; passwords are never returned."""
    data = _load_account_store()
    result: dict[str, dict] = {}
    for item in data.get('accounts', []):
        if not isinstance(item, dict):
            continue
        email = str(item.get('email', '') or '').strip()
        if not email:
            continue
        result[email.casefold()] = {
            'email': email,
            'source': 'app',
            'has_password': bool(_decode_password(str(item.get('password', '') or ''))),
        }
    if include_legacy and base is not None:
        try:
            legacy, _book, _selected = _legacy_logins(base)
        except Exception:
            legacy = []
        for email, password in legacy:
            key = email.casefold()
            if key not in result:
                result[key] = {'email': email, 'source': 'legacy workbook', 'has_password': bool(password)}
    return sorted(result.values(), key=lambda x: x['email'].casefold())


def save_shedsuite_account(email: str, password: str) -> bool:
    email = str(email or '').strip()
    password = str(password or '')
    if not email or '@' not in email or not password:
        return False
    data = _load_account_store()
    accounts = [x for x in data.get('accounts', []) if isinstance(x, dict)]
    encoded = _encode_password(password)
    found = False
    for item in accounts:
        if str(item.get('email', '') or '').strip().casefold() == email.casefold():
            item['email'] = email
            item['password'] = encoded
            found = True
            break
    if not found:
        accounts.append({'email': email, 'password': encoded})
    data['accounts'] = accounts
    data['version'] = 1
    _save_account_store(data)
    return True


def delete_shedsuite_account(email: str) -> bool:
    email = str(email or '').strip()
    if not email:
        return False
    data = _load_account_store()
    old = [x for x in data.get('accounts', []) if isinstance(x, dict)]
    new = [x for x in old if str(x.get('email', '') or '').strip().casefold() != email.casefold()]
    if len(new) == len(old):
        return False
    data['accounts'] = new
    _save_account_store(data)
    remove_shedsuite_login_mappings_for_email(email)
    return True


def _stored_logins() -> list[tuple[str, str]]:
    data = _load_account_store()
    result: list[tuple[str, str]] = []
    for item in data.get('accounts', []):
        if not isinstance(item, dict):
            continue
        email = str(item.get('email', '') or '').strip()
        password = _decode_password(str(item.get('password', '') or ''))
        if email and password:
            result.append((email, password))
    return result


def _all_logins(base: Path) -> tuple[list[tuple[str, str]], Path | None, str]:
    """Merge in-app accounts with the legacy Logininfo table; in-app wins."""
    legacy, book, selected = _legacy_logins(base)
    merged: dict[str, tuple[str, str]] = {email.casefold(): (email, password) for email, password in legacy}
    for email, password in _stored_logins():
        merged[email.casefold()] = (email, password)
    result = list(merged.values())
    if selected:
        result.sort(key=lambda x: 0 if x[0].casefold() == selected.casefold() else 1)
    return result, book, selected


def _login_mapping_path() -> Path:
    override = str(os.getenv('SHEDSUITE_LOGIN_MAPPING_FILE', '') or '').strip()
    if override:
        return Path(override).expanduser()
    if os.name == 'nt':
        root = Path(os.getenv('LOCALAPPDATA') or Path.home())
        return root / 'ShedSuiteRTO' / 'shedsuite_login_mappings.json'
    return Path.home() / '.shedsuite_rto' / 'shedsuite_login_mappings.json'


def _login_entity_key(value: str) -> str:
    text = str(value or '').casefold().replace('&', ' and ')
    words = re.findall(r'[a-z0-9]+', text)
    disposable = {
        'llc', 'inc', 'incorporated', 'corp', 'corporation', 'company', 'co',
        'limited', 'ltd', 'buildings', 'building', 'enterprises', 'enterprise',
    }
    words = [w for w in words if w not in disposable]
    return ' '.join(words).strip()


def shedsuite_login_key(company: str, dealer: str = '') -> str:
    """Stable company key used for local ShedSuite-login learning."""
    company_key = _login_entity_key(company)
    if company_key:
        return company_key
    return _login_entity_key(dealer)


def _load_login_mappings() -> dict:
    path = _login_mapping_path()
    with _LOGIN_MAPPING_LOCK:
        if not path.exists():
            return {'version': 1, 'mappings': {}}
        try:
            data = json.loads(path.read_text(encoding='utf-8'))
        except Exception:
            return {'version': 1, 'mappings': {}}
        if not isinstance(data, dict):
            return {'version': 1, 'mappings': {}}
        if not isinstance(data.get('mappings'), dict):
            data['mappings'] = {}
        return data


def _save_login_mappings(data: dict) -> None:
    path = _login_mapping_path()
    with _LOGIN_MAPPING_LOCK:
        path.parent.mkdir(parents=True, exist_ok=True)
        temp = path.with_suffix('.tmp')
        temp.write_text(json.dumps(data, indent=2), encoding='utf-8')
        temp.replace(path)


def saved_shedsuite_login(company: str, dealer: str = '') -> str:
    key = shedsuite_login_key(company, dealer)
    if not key:
        return ''
    item = (_load_login_mappings().get('mappings', {}) or {}).get(key, {})
    return str(item.get('email', '') if isinstance(item, dict) else item or '').strip()


def save_shedsuite_login_mapping(company: str, dealer: str, email: str) -> bool:
    key = shedsuite_login_key(company, dealer)
    email = str(email or '').strip()
    if not key or not email:
        return False
    data = _load_login_mappings()
    mappings = data.setdefault('mappings', {})
    mappings[key] = {
        'email': email,
        'company': str(company or '').strip(),
        'dealer': str(dealer or '').strip(),
    }
    data['version'] = 1
    _save_login_mappings(data)
    return True


def remove_shedsuite_login_mappings_for_email(email: str) -> None:
    email = str(email or '').strip().casefold()
    if not email:
        return
    data = _load_login_mappings()
    mappings = data.setdefault('mappings', {})
    changed = False
    for key in list(mappings):
        item = mappings.get(key, {})
        mapped = str(item.get('email', '') if isinstance(item, dict) else item or '').strip().casefold()
        if mapped == email:
            mappings.pop(key, None)
            changed = True
    if changed:
        _save_login_mappings(data)


def available_shedsuite_login_emails(base: Path) -> list[str]:
    """Return all locally available account emails without exposing passwords."""
    logins, _book, _selected = _all_logins(base)
    return [email for email, _password in logins]


def resolve_shedsuite_login(company: str, dealer: str = '', explicit_login: str = '') -> tuple[str, str]:
    """Resolve one deterministic login and report why it was chosen.

    This function deliberately never returns an arbitrary fallback login.  If a
    new company is not mapped, the UI must ask once instead of Chromium trying
    unrelated tenants for up to two minutes each.
    """
    explicit_login = str(explicit_login or '').strip()
    # A user-confirmed saved mapping is allowed to override an older built-in
    # company rule. That makes onboarding/fixing a login possible without
    # editing Python again.
    learned = saved_shedsuite_login(company, dealer)
    if learned:
        return learned, 'saved mapping'
    if explicit_login:
        return explicit_login, 'company rule'
    preferred = _preferred_login(company, dealer)
    if preferred:
        return preferred[0], 'built-in mapping'
    return '', 'mapping needed'


def _safe_email(email: str) -> str:
    return re.sub(r'[^a-zA-Z0-9._-]', '_', (email or '').strip().lower())


def _find_legacy_workbook(base: Path) -> Path | None:
    """Find the old workbook that already contains the Logininfo table.

    V5 intentionally does NOT ask for a ShedSuite password. It reuses the same
    credentials source as Contract_Import.xlsm / Contract_Import.py.
    """
    explicit = os.getenv('LEGACY_XLSM_PATH', '').strip()
    candidates: list[Path] = []
    if explicit:
        candidates.append(Path(explicit))

    fixed_roots = [base, base.parent]
    home = Path(os.environ.get('USERPROFILE') or Path.home())
    fixed_roots += [home / 'Downloads', home / 'Desktop', home / 'Documents']

    for root in fixed_roots:
        candidates += [
            root / 'Contract_Import.xlsm',
            root / 'Shedsuite_Import V1.1.7.6' / 'Contract_Import.xlsm',
            root / 'Shedsuite_Import V1.1.7.6' / 'Shedsuite_Import V1.1.7.6' / 'Contract_Import.xlsm',
        ]

    for p in candidates:
        try:
            if p.exists() and p.is_file():
                return p.resolve()
        except Exception:
            pass

    # Last-resort bounded search of the normal user folders and the current
    # app's parent tree. New V7 builds are often extracted into another nested
    # version folder, while Contract_Import.xlsm remains beside an older build.
    # Search several levels without ever scanning the whole drive.
    search_roots = [base.parent, base.parent.parent, home / 'Downloads', home / 'Desktop', home / 'Documents']
    seen_roots = set()
    for root in search_roots:
        try:
            root = root.resolve()
        except Exception:
            continue
        if root in seen_roots or not root.exists():
            continue
        seen_roots.add(root)
        try:
            root_depth = len(root.parts)
            for current, dirs, files in os.walk(root):
                current_path = Path(current)
                depth = len(current_path.parts) - root_depth
                # Four nested folders is enough for typical Desktop\shed\version
                # layouts while avoiding an expensive recursive user-profile scan.
                if depth >= 4:
                    dirs[:] = []
                if 'Contract_Import.xlsm' in files:
                    return (current_path / 'Contract_Import.xlsm').resolve()
        except Exception:
            pass
    return None


def _legacy_logins(base: Path) -> tuple[list[tuple[str, str]], Path | None, str]:
    """Read the exact Excel Logininfo table. Passwords remain local/in-memory."""
    book = _find_legacy_workbook(base)
    if not book:
        return [], None, ''

    wb = load_workbook(book, data_only=True, read_only=False, keep_vba=True)
    try:
        selected = str(wb['Import Tools']['D4'].value or '').strip() if 'Import Tools' in wb.sheetnames else ''
        if 'RTOPro Data' not in wb.sheetnames:
            return [], book, selected
        ws = wb['RTOPro Data']
        table = ws.tables.get('Logininfo')
        result: list[tuple[str, str]] = []
        if table:
            rows = list(ws[table.ref])
            for row in rows[1:]:
                email = str(row[0].value or '').strip()
                password = str(row[1].value or '')
                if email and password:
                    result.append((email, password))
    finally:
        wb.close()

    if selected:
        result.sort(key=lambda x: 0 if x[0].lower() == selected.lower() else 1)
    return result, book, selected


def _preferred_login(company: str, dealer: str) -> list[str]:
    """Same company/login selection idea used by the old Excel macro."""
    text = f'{company} {dealer}'.lower()
    if 'smart shed' in text or 'magnolia' in text:
        return ['office@magnoliarto.com']
    if 'lonestar' in text or re.search(r'\blsr\b', text):
        return ['lsrbarns@gmail.com']
    if 'westwood' in text or re.search(r'\bwwp\b', text):
        return ['office@westwoodpayments.com']
    if '4 seasons' in text or 'four seasons' in text:
        return ['conrad.s+wv@evergreenrto.com']
    if 'phoenix' in text and ('dbm' in text or 'dutch' in text):
        return ['conrad.s+dbm@evergreenrto.com']
    if 'yoder storage' in text or 'yss' in text:
        return ['conrad.s+ysb@evergreenrto.com', 'joy.f+davis@evergreenrto.com']
    if 'alpine' in text:
        return ['tori.h@evergreenrto.com', 'rentabarn2@gmail.com']
    if 'genesis' in text:
        return ['conrad.s@evergreenrto.com']
    return []


def _unwrap_url(url: str) -> str:
    """Legacy ShedSuite viewer URLs sometimes contain a second real https URL."""
    if not url:
        return ''
    hits = [m.start() for m in re.finditer('https', str(url))]
    if len(hits) >= 2:
        return str(url)[hits[1]:]
    if hits:
        return str(url)[hits[0]:]
    return str(url)


async def _url_to_pdf_bytes(context, url: str) -> bytes | None:
    url = _unwrap_url(url)
    if not url or url.startswith(('blob:', 'data:')):
        return None
    try:
        response = await context.request.get(url, timeout=60000)
        if not response.ok:
            return None
        return bytes_to_pdf(await response.body())
    except Exception:
        return None


async def _page_blob_bytes(page) -> bytes | None:
    try:
        data_url = await page.evaluate('''async () => {
            const response = await fetch(window.location.href);
            const blob = await response.blob();
            return await new Promise((resolve, reject) => {
                const reader = new FileReader();
                reader.onload = () => resolve(reader.result);
                reader.onerror = reject;
                reader.readAsDataURL(blob);
            });
        }''')
        if not data_url or ',' not in data_url:
            return None
        return bytes_to_pdf(base64.b64decode(data_url.split(',', 1)[1]))
    except Exception:
        return None


async def _login_context(browser, email: str, password: str, base: Path, legacy_book: Path | None):
    """Open ShedSuite without prompting; auto-login from old Excel if needed."""
    local_auth = base / 'auth_state' / f'shedsuite_auth_{_safe_email(email)}.json'
    local_auth.parent.mkdir(parents=True, exist_ok=True)
    legacy_auth = legacy_book.parent / 'auth_state' / local_auth.name if legacy_book else None
    auth = local_auth if local_auth.exists() else (legacy_auth if legacy_auth and legacy_auth.exists() else None)

    kwargs = {
        'viewport': {'width': 1100, 'height': 850},
        'accept_downloads': True,
        'user_agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        ),
    }
    if auth:
        kwargs['storage_state'] = str(auth)

    context = await browser.new_context(**kwargs)
    context.set_default_navigation_timeout(90000)
    context.set_default_timeout(15000)
    page = await context.new_page()
    await page.goto(HOME_URL, wait_until='domcontentloaded', timeout=90000)
    await page.wait_for_timeout(2500)

    login_button = page.get_by_role('button', name=re.compile(r'log\s*in', re.I))
    email_box = page.get_by_label(re.compile(r'email', re.I))
    password_box = page.get_by_label(re.compile(r'password', re.I))
    try:
        needs_login = await login_button.is_visible()
    except Exception:
        needs_login = False
    if not needs_login:
        try:
            needs_login = await email_box.is_visible() and await password_box.is_visible()
        except Exception:
            needs_login = False

    if needs_login:
        # No prompt: this is exactly why we read the old workbook's Logininfo.
        try:
            await email_box.fill(email)
        except Exception:
            await page.locator('input[type="email"]').first.fill(email)
        try:
            await password_box.fill(password)
        except Exception:
            await page.locator('input[type="password"]').first.fill(password)
        try:
            await login_button.click()
        except Exception:
            await page.locator('button[type="submit"]').first.click()
        try:
            await expect(page).to_have_url(re.compile(r'https://app\.shedsuite\.com/rto/.*'), timeout=300000)
        except Exception:
            # ShedSuite occasionally lands on another authenticated app route.
            # A successful disappearance of the login fields is sufficient.
            await page.wait_for_timeout(2500)
        await context.storage_state(path=str(local_auth))

    return context, page


async def _capture_delivery_certificate(page, context, order_id: str) -> tuple[bytes | None, str, str]:
    """Capture only the Delivery Certificate from a ShedSuite order.

    V7.6.2 deliberately waits for ShedSuite's React attachment list.  Earlier
    V7.6.x builds inspected the DOM immediately after ``domcontentloaded`` and
    could therefore report a false Missing result while the Files list was
    still rendering.
    """
    order_url = BASE_URL + str(order_id)
    responses = []

    def remember(response):
        responses.append(response)

    async def login_screen_visible() -> bool:
        try:
            if '/login' in str(page.url).lower():
                return True
        except Exception:
            pass
        try:
            email = page.locator('input[type="email"]')
            password = page.locator('input[type="password"]')
            return (await email.count() > 0 and await password.count() > 0
                    and await email.first.is_visible() and await password.first.is_visible())
        except Exception:
            return False

    async def locate_delivery_target():
        candidate_selectors = [
            'span[class^="fileName"]',
            'span[class*="fileName"]',
            '[class*="fileName"]',
        ]
        for selector in candidate_selectors:
            try:
                locs = page.locator(selector)
                count = await locs.count()
                for i in range(count):
                    loc = locs.nth(i)
                    name = ((await loc.text_content()) or '').strip()
                    low = re.sub(r'\s+', ' ', name.lower())
                    if 'delivery' in low and ('certificate' in low or re.search(r'\bcert\b', low)):
                        return loc, name
            except Exception:
                continue

        # Fall back to text rather than generated CSS classes.  Exact=False is
        # intentional because ShedSuite sometimes adds a date/file extension.
        try:
            text_matches = page.get_by_text(
                re.compile(r'delivery\s*(certificate|cert(?:ificate)?)', re.I)
            )
            count = await text_matches.count()
            for i in range(count):
                loc = text_matches.nth(i)
                try:
                    if not await loc.is_visible():
                        continue
                except Exception:
                    pass
                name = ((await loc.text_content()) or '').strip()
                if name:
                    return loc, name
        except Exception:
            pass
        return None, ''

    async def wait_for_delivery_target(timeout_ms: int):
        """Poll because ShedSuite's Files section is populated after page load."""
        loop = asyncio.get_running_loop()
        deadline = loop.time() + (timeout_ms / 1000.0)
        files_seen_at = None
        last_body = ''
        while loop.time() < deadline:
            if await login_screen_visible():
                return None, '', 'ShedSuite session is not authenticated (login page was shown).'

            target, name = await locate_delivery_target()
            if target is not None:
                return target, name, ''

            # Once the attachment list itself is clearly present, allow a few
            # extra seconds for all rows to render before deciding the certificate
            # truly is absent.
            try:
                file_count = await page.locator('[class*="fileName"]').count()
            except Exception:
                file_count = 0
            if file_count:
                if files_seen_at is None:
                    files_seen_at = loop.time()
                elif loop.time() - files_seen_at >= 8:
                    try:
                        last_body = (await page.locator('body').inner_text()).lower()
                    except Exception:
                        last_body = ''
                    if 'delivery certificate' not in last_body and 'delivery cert' not in last_body:
                        return None, '', 'Delivery Certificate was not listed on the ShedSuite order page.'

            await page.wait_for_timeout(1000)

        try:
            last_body = (await page.locator('body').inner_text()).lower()
        except Exception:
            last_body = ''
        if 'delivery certificate' in last_body or 'delivery cert' in last_body:
            return None, '', 'Delivery Certificate text was present but its clickable element could not be located.'
        return None, '', 'Timed out waiting for ShedSuite to load the Delivery Certificate/file list.'

    try:
        # Try twice. A soft reload fixes occasional ShedSuite orders whose Files
        # request stalls even though the order page itself has loaded.
        target = None
        found_name = ''
        find_error = ''
        for attempt in range(2):
            if attempt == 0:
                await page.goto(order_url, wait_until='domcontentloaded', timeout=90000)
            else:
                try:
                    await page.reload(wait_until='domcontentloaded', timeout=90000)
                except Exception:
                    await page.goto(order_url, wait_until='domcontentloaded', timeout=90000)

            # Keep the historical 90-second tolerance on the first attempt. The
            # retry is shorter because it is only a recovery path.
            target, found_name, find_error = await wait_for_delivery_target(90000 if attempt == 0 else 30000)
            if target is not None:
                break
            # Authentication problems will not be repaired by refreshing.
            if 'not authenticated' in find_error.lower():
                return None, find_error, ''

        if target is None:
            return None, find_error or 'Delivery Certificate not found on the ShedSuite order page.', ''

        # Some ShedSuite builds expose an authenticated anchor directly.
        try:
            href = await target.evaluate("el => (el.closest('a') && el.closest('a').href) || ''")
            if href:
                data = await _url_to_pdf_bytes(context, href)
                if data:
                    return data, 'authenticated link', found_name
        except Exception:
            pass

        page.on('response', remember)
        popup_task = asyncio.create_task(page.wait_for_event('popup'))
        download_task = asyncio.create_task(page.wait_for_event('download'))
        before_url = page.url

        try:
            # Prefer the nearest clickable ancestor. Clicking a nested text span
            # does not always trigger the document viewer in newer ShedSuite UI.
            try:
                clickable = target.locator('xpath=ancestor-or-self::a[1] | ancestor-or-self::button[1] | ancestor-or-self::*[@role="button"][1]').first
                if await clickable.count() and await clickable.is_visible():
                    await clickable.click(timeout=15000)
                else:
                    await target.click(timeout=15000)
            except Exception:
                await target.click(timeout=15000)
        except Exception as exc:
            for task in (popup_task, download_task):
                if not task.done():
                    task.cancel()
            return None, f'click failed: {exc}', found_name

        done, pending = await asyncio.wait(
            [popup_task, download_task], timeout=20, return_when=asyncio.FIRST_COMPLETED
        )
        for task in pending:
            task.cancel()

        if download_task in done:
            try:
                download = await download_task
                with tempfile.NamedTemporaryFile(delete=False, suffix='.bin') as tf:
                    temp_path = Path(tf.name)
                try:
                    await download.save_as(str(temp_path))
                    raw = temp_path.read_bytes()
                    converted = bytes_to_pdf(raw)
                    if converted:
                        return converted, 'browser download', found_name
                finally:
                    temp_path.unlink(missing_ok=True)
            except Exception:
                pass

        if popup_task in done:
            try:
                popup = await popup_task
                try:
                    await popup.wait_for_load_state('domcontentloaded', timeout=30000)
                except Exception:
                    pass
                data = None
                if popup.url.startswith(('blob:', 'data:')):
                    data = await _page_blob_bytes(popup)
                else:
                    data = await _url_to_pdf_bytes(context, popup.url)
                    if not data:
                        data = await _page_blob_bytes(popup)
                try:
                    await popup.close()
                except Exception:
                    pass
                if data:
                    return data, 'popup/viewer', found_name
            except Exception:
                pass

        # A PDF viewer may fetch the bytes in the background without a useful URL.
        await page.wait_for_timeout(3000)
        for response in reversed(responses):
            try:
                ctype = str((response.headers or {}).get('content-type', '')).lower()
                url = str(response.url or '').lower()
                if 'pdf' in ctype or 'application/octet-stream' in ctype or 'delivery' in url or 'certificate' in url:
                    raw = await response.body()
                    if raw:
                        converted = bytes_to_pdf(raw)
                        if converted:
                            return converted, 'network response', found_name
            except Exception:
                continue

        # Final fallback: the document replaced the order page in the same tab.
        if page.url != before_url:
            try:
                data = await (_page_blob_bytes(page) if page.url.startswith(('blob:', 'data:')) else _url_to_pdf_bytes(context, page.url))
                if not data:
                    data = await _page_blob_bytes(page)
                if data:
                    return data, 'same-tab viewer', found_name
            except Exception:
                pass

        return None, 'Delivery Certificate was found and clicked, but no PDF bytes were captured.', found_name
    except Exception as exc:
        return None, str(exc), ''
    finally:
        try:
            page.remove_listener('response', remember)
        except Exception:
            pass

def _append_pdf(target: Path, cert_bytes: bytes) -> None:
    cert = fitz.open(stream=cert_bytes, filetype='pdf')
    try:
        if target.exists():
            current = fitz.open(target)
            try:
                current.insert_pdf(cert)
                temp = target.with_suffix('.deliverytmp.pdf')
                current.save(temp, garbage=4, deflate=True)
            finally:
                current.close()
            temp.replace(target)
        else:
            cert.save(target, garbage=4, deflate=True)
    finally:
        cert.close()


async def _run(
    rows: list[dict],
    source_rows: list[dict],
    pdf_dir: Path,
    base: Path,
    progress=None,
    should_process=None,
    headless: bool = True,
    prefetch_cache_dir: Path | None = None,
    packet_ready_flag: Path | None = None,
    concurrency: int = 3,
) -> list[str]:
    """Download Delivery Certificates with session reuse and bounded concurrency.

    V7.15 can start this before the direct ShedSuite packet PDFs are finished.
    Certificates are fetched into a small local cache first; once the base-packet
    marker appears they are appended to the combined PDFs.  This lets Chromium do
    network/login work while the normal /process request is still assembling the
    faster direct-link documents.

    One authenticated context/page is reused per ShedSuite login.  A per-login
    asyncio lock keeps that page serial, while different company logins can run
    concurrently (default: 3 accounts at once).
    """
    logins, legacy_book, _selected = _all_logins(base)
    if prefetch_cache_dir:
        prefetch_cache_dir.mkdir(parents=True, exist_ok=True)

    async def wait_for_packet_ready(timeout_seconds: float = 600.0) -> bool:
        if packet_ready_flag is None:
            return True
        loop = asyncio.get_running_loop()
        deadline = loop.time() + timeout_seconds
        while loop.time() < deadline:
            if packet_ready_flag.exists():
                return True
            await asyncio.sleep(0.25)
        return packet_ready_flag.exists()

    if not logins:
        msg = (
            'Delivery Certificate: no ShedSuite accounts are saved. '
            'Open ShedSuite Accounts in the web app and add at least one login.'
        )
        for row in rows:
            oid = str(row.get('_source_order_id', '')).strip()
            if oid:
                row['_delivery_certificate'] = 'Login mapping needed'
                row['_delivery_certificate_method'] = ''
                if progress:
                    progress(oid, 'login_needed', 'No ShedSuite accounts are saved. Add one in ShedSuite Accounts.', row)
        # In prefetch mode, do not let the app merge/rebuild while /process is
        # still constructing the base packets.
        await wait_for_packet_ready()
        return [msg]

    src_by_order = {str(x.get('Customer Order Id', '')).strip(): x for x in source_rows}
    login_map = {email.lower(): (email, password) for email, password in logins}
    warnings: list[str] = []

    try:
        concurrency = max(1, min(5, int(concurrency or 3)))
    except Exception:
        concurrency = 3

    async with async_playwright() as playwright:
        launch_kwargs = {'headless': headless}
        if not headless:
            launch_kwargs['args'] = [
                '--window-position=-32000,-32000',
                '--window-size=1200,900',
                '--disable-background-timer-throttling',
                '--disable-renderer-backgrounding',
                '--disable-backgrounding-occluded-windows',
                '--disable-features=CalculateNativeWinOcclusion',
            ]
        browser = await playwright.chromium.launch(**launch_kwargs)
        contexts: dict[str, tuple] = {}
        login_locks: dict[str, asyncio.Lock] = {}
        semaphore = asyncio.Semaphore(concurrency)
        pending_prefetch: dict[str, tuple[dict, Path, str, str]] = {}

        def login_resolution_for(row: dict) -> tuple[tuple[str, str] | None, str, str, str, str]:
            """Resolve a known login; unknown companies are auto-discovered."""
            oid = str(row.get('_source_order_id', '')).strip()
            src = src_by_order.get(oid, {})
            company = str(src.get('Company Name', '') or row.get('_source_company', '') or '').strip()
            dealer = str(src.get('Dealer', '') or row.get('_source_dealer', '') or '').strip()
            explicit = str(row.get('_login') or '').strip()
            email, reason = resolve_shedsuite_login(company, dealer, explicit)
            if not email:
                return None, '', reason, company, dealer
            item = login_map.get(email.lower())
            if not item:
                return None, email, 'mapped login is not present in saved accounts', company, dealer
            return item, email, reason, company, dealer

        discovery_cache: dict[str, str] = {}
        discovery_locks: dict[str, asyncio.Lock] = {}
        validated_mapping_cache: dict[str, str] = {}

        async def probe_account_for_order(email: str, password: str, oid: str, row: dict) -> bool:
            """Quickly determine whether one tenant owns the order without waiting for attachments."""
            key = email.lower()
            lock = login_locks.setdefault(key, asyncio.Lock())
            async with lock:
                context, page = await get_context(email, password)
                try:
                    await page.goto(BASE_URL + str(oid), wait_until='domcontentloaded', timeout=90000)
                except Exception:
                    return False
                await page.wait_for_timeout(1800)
                try:
                    if '/login' in str(page.url).lower():
                        return False
                except Exception:
                    pass
                try:
                    body = re.sub(r'\s+', ' ', (await page.locator('body').inner_text()) or '').casefold()
                except Exception:
                    body = ''
                negatives = ('order not found', 'not found', 'does not exist', 'access denied', 'unauthorized', 'forbidden')
                if any(x in body for x in negatives):
                    return False
                if str(oid).casefold() in body:
                    return True
                name = str(row.get('NAME', '') or '').strip().casefold()
                last = name.split()[-1] if name else ''
                try:
                    file_count = await page.locator('[class*="fileName"]').count()
                except Exception:
                    file_count = 0
                positive_markers = ('customer information', 'order details', 'contract information', 'files', 'documents')
                if file_count > 0:
                    return True
                if last and len(last) >= 3 and last in body and any(x in body for x in positive_markers):
                    return True
                return False

        async def discover_login_for_row(row: dict, oid: str, company: str, dealer: str) -> tuple[tuple[str, str] | None, str, str]:
            """Try saved accounts automatically and remember the tenant that owns this company/order."""
            entity_key = shedsuite_login_key(company, dealer) or f'order:{oid}'
            if entity_key in discovery_cache:
                email = discovery_cache[entity_key]
                item = login_map.get(email.casefold())
                if item:
                    return item, email, 'auto-discovered mapping'

            lock = discovery_locks.setdefault(entity_key, asyncio.Lock())
            async with lock:
                if entity_key in discovery_cache:
                    email = discovery_cache[entity_key]
                    item = login_map.get(email.casefold())
                    if item:
                        return item, email, 'auto-discovered mapping'

                candidates = list(logins)
                if not candidates:
                    return None, '', 'no saved ShedSuite accounts'
                if progress:
                    progress(oid, 'working', f'New company: automatically checking {len(candidates)} saved ShedSuite account(s)…', row)

                # Probe in bounded parallel batches. We reuse authenticated contexts,
                # so later discoveries are usually much faster than the first one.
                async def one(candidate):
                    email, password = candidate
                    try:
                        ok = await probe_account_for_order(email, password, oid, row)
                    except Exception:
                        ok = False
                    return candidate if ok else None

                # Use the same global semaphore to avoid flooding ShedSuite.
                async def guarded(candidate):
                    async with semaphore:
                        return await one(candidate)

                results = await asyncio.gather(*(guarded(c) for c in candidates))
                matches = [x for x in results if x]
                if len(matches) == 1:
                    email, password = matches[0]
                    discovery_cache[entity_key] = email
                    save_shedsuite_login_mapping(company, dealer, email)
                    row['_login'] = email
                    if progress:
                        progress(oid, 'working', f'Automatically discovered ShedSuite login {email}. Downloading certificate…', row)
                    return (email, password), email, 'auto-discovered mapping'
                if len(matches) > 1:
                    return None, '', 'multiple saved ShedSuite accounts appeared to contain this order; choose one manually'
                return None, '', 'automatic login discovery could not find this order in the saved ShedSuite accounts'

        async def get_context(email: str, password: str):
            key = email.lower()
            if key not in contexts:
                contexts[key] = await _login_context(browser, email, password, base, legacy_book)
            return contexts[key]

        async def append_prefetched(row: dict, cache_path: Path, method: str, cert_name: str) -> bool:
            oid = str(row.get('_source_order_id', '')).strip()
            if should_process and not should_process(oid):
                try:
                    cache_path.unlink(missing_ok=True)
                except Exception:
                    pass
                return False
            try:
                cert_data = cache_path.read_bytes()
                target = pdf_dir / f"{safe_filename(row.get('_pdf_name') or row.get('NAME') or oid)}.pdf"
                _append_pdf(target, cert_data)
                try:
                    cache_path.unlink(missing_ok=True)
                except Exception:
                    pass
                row['_delivery_certificate'] = 'Downloaded'
                row['_delivery_certificate_method'] = method
                row['_delivery_certificate_name'] = cert_name or 'Delivery Certificate'
                if progress:
                    progress(oid, 'ready', f'Downloaded via {method}.', row)
                return True
            except Exception as exc:
                row['_delivery_certificate'] = 'Missing'
                row['_delivery_certificate_method'] = ''
                warnings.append(f"{row.get('NAME') or oid}: Delivery Certificate append failed: {exc}")
                if progress:
                    progress(oid, 'missing', f'Certificate was prefetched but could not be appended: {exc}', row)
                return False

        async def process_row(row: dict) -> None:
            oid = str(row.get('_source_order_id', '')).strip()
            if not oid:
                return
            if should_process and not should_process(oid):
                return

            known_item, resolved_email, login_reason, company, dealer = login_resolution_for(row)
            entity_key = shedsuite_login_key(company, dealer) or f'order:{oid}'

            # V7.22 validates a known mapping once per company before trusting it.
            # If an older built-in/saved mapping points to the wrong tenant, the
            # app automatically falls back to account discovery instead of waiting
            # on the wrong order page or making the user close Chromium manually.
            if known_item is not None:
                cached_email = validated_mapping_cache.get(entity_key, '')
                if cached_email.casefold() != resolved_email.casefold():
                    try:
                        async with semaphore:
                            owns_order = await probe_account_for_order(known_item[0], known_item[1], oid, row)
                    except Exception:
                        owns_order = False
                    if owns_order:
                        validated_mapping_cache[entity_key] = resolved_email
                    else:
                        if progress:
                            progress(oid, 'working', f'Mapped login {resolved_email} did not match this order. Auto-discovering the correct ShedSuite account…', row)
                        known_item, resolved_email, login_reason = await discover_login_for_row(row, oid, company, dealer)
                        if known_item is not None:
                            validated_mapping_cache[entity_key] = resolved_email

            if known_item is None:
                prior_resolved_email = resolved_email
                discovered_item, discovered_email, discovered_reason = await discover_login_for_row(row, oid, company, dealer)
                if discovered_item is not None:
                    known_item, resolved_email, login_reason = discovered_item, discovered_email, discovered_reason
                    validated_mapping_cache[entity_key] = resolved_email
                elif not resolved_email:
                    resolved_email = prior_resolved_email
                    login_reason = discovered_reason or login_reason

            if known_item is None:
                row['_delivery_certificate'] = 'Login mapping needed'
                row['_delivery_certificate_method'] = ''
                if resolved_email:
                    detail = f'ShedSuite login {resolved_email} is mapped to this company but is not present in the saved accounts.'
                else:
                    detail = login_reason or f'ShedSuite login mapping needed for {company or dealer or "this company"}.'
                warnings.append(f"{row.get('NAME') or oid}: Delivery Certificate: {detail}")
                if progress:
                    progress(oid, 'login_needed', detail, row)
                return

            if progress:
                progress(oid, 'working', f'Using ShedSuite login {resolved_email} ({login_reason}). Prefetching Delivery Certificate…', row)

            cert_data = None
            method = ''
            cert_name = ''
            last_error = 'Delivery Certificate not found'

            email, password = known_item
            async with semaphore:
                if should_process and not should_process(oid):
                    return
                key = email.lower()
                lock = login_locks.setdefault(key, asyncio.Lock())
                try:
                    async with lock:
                        context, page = await get_context(email, password)
                        cert_data, method, cert_name = await _capture_delivery_certificate(page, context, oid)
                    if not cert_data:
                        last_error = method
                except Exception as exc:
                    last_error = str(exc)

            if cert_data and resolved_email:
                src = src_by_order.get(oid, {})
                save_shedsuite_login_mapping(
                    str(src.get('Company Name', '') or row.get('_source_company', '') or '').strip(),
                    str(src.get('Dealer', '') or row.get('_source_dealer', '') or '').strip(),
                    resolved_email,
                )

            if not cert_data:
                row['_delivery_certificate'] = 'Missing'
                row['_delivery_certificate_method'] = ''
                warnings.append(f"{row.get('NAME') or oid}: Delivery Certificate: {last_error}")
                if progress:
                    progress(oid, 'missing', last_error, row)
                return

            if prefetch_cache_dir is None:
                target = pdf_dir / f"{safe_filename(row.get('_pdf_name') or row.get('NAME') or oid)}.pdf"
                _append_pdf(target, cert_data)
                row['_delivery_certificate'] = 'Downloaded'
                row['_delivery_certificate_method'] = method
                row['_delivery_certificate_name'] = cert_name or 'Delivery Certificate'
                if progress:
                    progress(oid, 'ready', f'Downloaded via {method}.', row)
                return

            cache_path = prefetch_cache_dir / f'{safe_filename(oid)}.pdf'
            temp = cache_path.with_suffix('.tmp')
            temp.write_bytes(cert_data)
            temp.replace(cache_path)
            row['_delivery_certificate'] = 'Prefetched'
            row['_delivery_certificate_method'] = method
            row['_delivery_certificate_name'] = cert_name or 'Delivery Certificate'

            # If the direct packet has already finished, append now. Otherwise
            # keep fetching later contracts rather than blocking this login on a
            # file-system wait; the second phase appends all cached certificates.
            if packet_ready_flag is None or packet_ready_flag.exists():
                await append_prefetched(row, cache_path, method, cert_name)
            else:
                pending_prefetch[oid] = (row, cache_path, method, cert_name)
                if progress:
                    progress(oid, 'prefetched', 'Certificate prefetched; waiting for the base contract packet…', row)

        try:
            tasks = [asyncio.create_task(process_row(row)) for row in rows]
            if tasks:
                await asyncio.gather(*tasks)

            if pending_prefetch:
                ready = await wait_for_packet_ready()
                if not ready:
                    for oid, (row, cache_path, _method, _cert_name) in list(pending_prefetch.items()):
                        if should_process and not should_process(oid):
                            continue
                        row['_delivery_certificate'] = 'Missing'
                        row['_delivery_certificate_method'] = ''
                        msg = 'Delivery Certificate was prefetched, but the base packet never became ready.'
                        warnings.append(f"{row.get('NAME') or oid}: {msg}")
                        if progress:
                            progress(oid, 'missing', msg, row)
                else:
                    for oid, (row, cache_path, method, cert_name) in list(pending_prefetch.items()):
                        await append_prefetched(row, cache_path, method, cert_name)

            for email_key, (context, _page) in contexts.items():
                try:
                    await context.storage_state(path=str(base / 'auth_state' / f'shedsuite_auth_{_safe_email(email_key)}.json'))
                except Exception:
                    pass
        finally:
            for context, _page in contexts.values():
                try:
                    await context.close()
                except Exception:
                    pass
            await browser.close()

    return warnings

def append_delivery_certificates(
    rows: list[dict],
    source_rows: list[dict],
    pdf_dir: Path,
    base: Path,
    progress=None,
    should_process=None,
    headless: bool = True,
    prefetch_cache_dir: Path | None = None,
    packet_ready_flag: Path | None = None,
    concurrency: int = 3,
) -> list[str]:
    """Synchronous wrapper used by the V7 background worker."""
    return asyncio.run(_run(
        rows, source_rows, pdf_dir, base,
        progress=progress, should_process=should_process, headless=headless,
        prefetch_cache_dir=prefetch_cache_dir, packet_ready_flag=packet_ready_flag,
        concurrency=concurrency,
    ))
