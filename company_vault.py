from __future__ import annotations

import base64
import json
import os
import re
import threading
import uuid
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.scrypt import Scrypt


VAULT_FORMAT = 'ShedSuiteRTOVault'
VAULT_VERSION = 1
_VAULT_LOCK = threading.RLock()
_VAULT_DATA: dict | None = None
_VAULT_KEY: bytes | None = None
_VAULT_SALT: bytes | None = None


def vault_path() -> Path:
    override = str(os.getenv('SHEDSUITE_VAULT_FILE', '') or '').strip()
    if override:
        return Path(override).expanduser()
    return Path.home() / '.shedsuite_rto' / 'company_vault.enc'


def auth_state_dir() -> Path:
    override = str(os.getenv('SHEDSUITE_AUTH_DIR', '') or '').strip()
    if override:
        return Path(override).expanduser()
    return Path.home() / '.shedsuite_rto' / 'auth_state'


def _blank_vault() -> dict:
    return {
        'version': VAULT_VERSION,
        'accounts': [],
        'companies': [],
    }


def vault_exists() -> bool:
    return vault_path().exists()


def vault_unlocked() -> bool:
    return _VAULT_DATA is not None and _VAULT_KEY is not None and _VAULT_SALT is not None


def vault_status() -> dict:
    data = _VAULT_DATA if vault_unlocked() else None
    return {
        'exists': vault_exists(),
        'unlocked': vault_unlocked(),
        'path': str(vault_path()),
        'accounts': len(data.get('accounts', [])) if data else 0,
        'companies': len(data.get('companies', [])) if data else 0,
    }


def _derive_key(password: str, salt: bytes) -> bytes:
    password = str(password or '')
    if not password:
        raise ValueError('Master password is required.')
    kdf = Scrypt(salt=salt, length=32, n=2**15, r=8, p=1)
    return kdf.derive(password.encode('utf-8'))


def _encrypt_payload(data: dict, key: bytes, salt: bytes) -> dict:
    nonce = os.urandom(12)
    plain = json.dumps(data, ensure_ascii=False, separators=(',', ':')).encode('utf-8')
    aad = f'{VAULT_FORMAT}:{VAULT_VERSION}'.encode('ascii')
    ciphertext = AESGCM(key).encrypt(nonce, plain, aad)
    return {
        'format': VAULT_FORMAT,
        'version': VAULT_VERSION,
        'kdf': 'scrypt-32768-8-1',
        'salt': base64.b64encode(salt).decode('ascii'),
        'nonce': base64.b64encode(nonce).decode('ascii'),
        'ciphertext': base64.b64encode(ciphertext).decode('ascii'),
    }


def _write_envelope(envelope: dict) -> None:
    path = vault_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_suffix(path.suffix + '.tmp')
    temp.write_text(json.dumps(envelope, indent=2), encoding='utf-8')
    try:
        os.chmod(temp, 0o600)
    except Exception:
        pass
    temp.replace(path)
    try:
        os.chmod(path, 0o600)
    except Exception:
        pass


def _save_unlocked() -> None:
    if not vault_unlocked():
        raise RuntimeError('Credential vault is locked.')
    assert _VAULT_DATA is not None and _VAULT_KEY is not None and _VAULT_SALT is not None
    _write_envelope(_encrypt_payload(_VAULT_DATA, _VAULT_KEY, _VAULT_SALT))


def create_vault(master_password: str) -> None:
    global _VAULT_DATA, _VAULT_KEY, _VAULT_SALT
    master_password = str(master_password or '')
    if len(master_password) < 8:
        raise ValueError('Use a master password with at least 8 characters.')
    with _VAULT_LOCK:
        if vault_exists():
            raise ValueError('An encrypted vault already exists. Unlock it instead.')
        salt = os.urandom(16)
        key = _derive_key(master_password, salt)
        data = _blank_vault()
        _write_envelope(_encrypt_payload(data, key, salt))
        _VAULT_DATA, _VAULT_KEY, _VAULT_SALT = data, key, salt


def unlock_vault(master_password: str) -> None:
    global _VAULT_DATA, _VAULT_KEY, _VAULT_SALT
    with _VAULT_LOCK:
        path = vault_path()
        if not path.exists():
            raise ValueError('No encrypted vault exists yet. Create it first.')
        try:
            envelope = json.loads(path.read_text(encoding='utf-8'))
            if envelope.get('format') != VAULT_FORMAT:
                raise ValueError('This is not a ShedSuite RTO credential vault.')
            salt = base64.b64decode(envelope['salt'])
            nonce = base64.b64decode(envelope['nonce'])
            ciphertext = base64.b64decode(envelope['ciphertext'])
            key = _derive_key(master_password, salt)
            aad = f'{VAULT_FORMAT}:{VAULT_VERSION}'.encode('ascii')
            plain = AESGCM(key).decrypt(nonce, ciphertext, aad)
            data = json.loads(plain.decode('utf-8'))
            if not isinstance(data, dict):
                raise ValueError('Vault contents are invalid.')
            data.setdefault('accounts', [])
            data.setdefault('companies', [])
        except Exception as exc:
            _VAULT_DATA = _VAULT_KEY = _VAULT_SALT = None
            raise ValueError('Could not unlock the vault. Check the master password.') from exc
        _VAULT_DATA, _VAULT_KEY, _VAULT_SALT = data, key, salt


def lock_vault() -> None:
    global _VAULT_DATA, _VAULT_KEY, _VAULT_SALT
    with _VAULT_LOCK:
        _VAULT_DATA = None
        _VAULT_KEY = None
        _VAULT_SALT = None


def change_master_password(new_password: str) -> None:
    global _VAULT_KEY, _VAULT_SALT
    new_password = str(new_password or '')
    if len(new_password) < 8:
        raise ValueError('Use a master password with at least 8 characters.')
    with _VAULT_LOCK:
        if not vault_unlocked():
            raise RuntimeError('Unlock the vault first.')
        salt = os.urandom(16)
        key = _derive_key(new_password, salt)
        assert _VAULT_DATA is not None
        _write_envelope(_encrypt_payload(_VAULT_DATA, key, salt))
        _VAULT_KEY, _VAULT_SALT = key, salt


def _data() -> dict:
    if not vault_unlocked() or _VAULT_DATA is None:
        raise RuntimeError('Credential vault is locked.')
    return _VAULT_DATA


def _new_id(prefix: str) -> str:
    return f'{prefix}_{uuid.uuid4().hex[:12]}'


def list_accounts_safe() -> list[dict]:
    with _VAULT_LOCK:
        if not vault_unlocked():
            return []
        result = []
        for item in _data().get('accounts', []):
            if not isinstance(item, dict):
                continue
            result.append({
                'id': str(item.get('id', '')),
                'label': str(item.get('label', '') or ''),
                'email': str(item.get('email', '') or ''),
                'has_password': bool(item.get('password')),
                'has_session': bool(item.get('storage_state')),
            })
        return sorted(result, key=lambda x: ((x['label'] or x['email']).casefold(), x['email'].casefold()))


def all_account_credentials() -> list[tuple[str, str, str]]:
    """Return (account_id, email, password) while unlocked. Never expose this to templates."""
    with _VAULT_LOCK:
        if not vault_unlocked():
            return []
        result = []
        for item in _data().get('accounts', []):
            if not isinstance(item, dict):
                continue
            aid = str(item.get('id', '') or '').strip()
            email = str(item.get('email', '') or '').strip()
            password = str(item.get('password', '') or '')
            if aid and email and password:
                result.append((aid, email, password))
        return result


def account_by_id(account_id: str, include_password: bool = False) -> dict | None:
    with _VAULT_LOCK:
        if not vault_unlocked():
            return None
        account_id = str(account_id or '').strip()
        for item in _data().get('accounts', []):
            if str(item.get('id', '') or '').strip() == account_id:
                result = dict(item)
                if not include_password:
                    result.pop('password', None)
                    result['has_password'] = bool(item.get('password'))
                return result
    return None


def account_by_email(email: str, include_password: bool = False) -> dict | None:
    email = str(email or '').strip().casefold()
    if not email:
        return None
    with _VAULT_LOCK:
        if not vault_unlocked():
            return None
        for item in _data().get('accounts', []):
            if str(item.get('email', '') or '').strip().casefold() == email:
                result = dict(item)
                if not include_password:
                    result.pop('password', None)
                    result['has_password'] = bool(item.get('password'))
                return result
    return None


def save_account(account_id: str, label: str, email: str, password: str = '') -> str:
    email = str(email or '').strip()
    if '@' not in email or '.' not in email.rsplit('@', 1)[-1]:
        raise ValueError('Enter a valid ShedSuite login email.')
    label = str(label or '').strip()
    password = str(password or '')
    with _VAULT_LOCK:
        data = _data()
        accounts = [x for x in data.get('accounts', []) if isinstance(x, dict)]
        existing = None
        if account_id:
            existing = next((x for x in accounts if str(x.get('id', '')) == str(account_id)), None)
        duplicate = next((x for x in accounts if str(x.get('email', '')).casefold() == email.casefold() and x is not existing), None)
        if duplicate:
            raise ValueError('That ShedSuite email is already saved.')
        password_changed = bool(password)
        if existing is None:
            if not password:
                raise ValueError('Password is required for a new account.')
            existing = {'id': _new_id('acct')}
            accounts.append(existing)
        elif not password:
            password = str(existing.get('password', '') or '')
        existing.update({'label': label, 'email': email, 'password': password})
        if password_changed:
            existing.pop('storage_state', None)
        data['accounts'] = accounts
        _save_unlocked()
        return str(existing['id'])


def account_storage_state(email: str) -> dict | None:
    account = account_by_email(email, include_password=True)
    state = (account or {}).get('storage_state')
    return dict(state) if isinstance(state, dict) else None


def save_account_storage_state(email: str, state: dict) -> bool:
    if not isinstance(state, dict):
        return False
    email_key = str(email or '').strip().casefold()
    if not email_key:
        return False
    with _VAULT_LOCK:
        data = _data()
        for item in data.get('accounts', []):
            if isinstance(item, dict) and str(item.get('email', '') or '').strip().casefold() == email_key:
                item['storage_state'] = state
                _save_unlocked()
                return True
    return False


def clear_account_storage_state(email: str) -> bool:
    email_key = str(email or '').strip().casefold()
    if not email_key or not vault_unlocked():
        return False
    with _VAULT_LOCK:
        data = _data()
        for item in data.get('accounts', []):
            if isinstance(item, dict) and str(item.get('email', '') or '').strip().casefold() == email_key:
                if 'storage_state' in item:
                    item.pop('storage_state', None)
                    _save_unlocked()
                return True
    return False


def delete_account(account_id: str) -> bool:
    with _VAULT_LOCK:
        data = _data()
        old = [x for x in data.get('accounts', []) if isinstance(x, dict)]
        new = [x for x in old if str(x.get('id', '')) != str(account_id)]
        if len(new) == len(old):
            return False
        data['accounts'] = new
        for company in data.get('companies', []):
            if isinstance(company, dict) and str(company.get('account_id', '')) == str(account_id):
                company['account_id'] = ''
        _save_unlocked()
        return True


def _normalize_entity(value: Any) -> str:
    text = str(value or '').casefold().replace('&', ' and ')
    words = re.findall(r'[a-z0-9]+', text)
    disposable = {
        'llc', 'inc', 'incorporated', 'corp', 'corporation', 'company', 'co',
        'limited', 'ltd', 'buildings', 'building', 'enterprises', 'enterprise',
    }
    words = [w for w in words if w not in disposable]
    return ' '.join(words).strip()


def _company_aliases(item: dict) -> list[str]:
    aliases = [str(item.get('name', '') or '')]
    raw = item.get('aliases', [])
    if isinstance(raw, str):
        raw = re.split(r'[,\n;]+', raw)
    if isinstance(raw, list):
        aliases.extend(str(x or '') for x in raw)
    return [x.strip() for x in aliases if x and x.strip()]


def list_companies_safe() -> list[dict]:
    with _VAULT_LOCK:
        if not vault_unlocked():
            return []
        accounts = {str(x.get('id', '')): x for x in _data().get('accounts', []) if isinstance(x, dict)}
        result = []
        for item in _data().get('companies', []):
            if not isinstance(item, dict):
                continue
            row = dict(item)
            row['aliases'] = _company_aliases(item)[1:]
            acct = accounts.get(str(item.get('account_id', '') or ''), {})
            row['account_email'] = str(acct.get('email', '') or '')
            row['account_label'] = str(acct.get('label', '') or '')
            result.append(row)
        return sorted(result, key=lambda x: str(x.get('name', '')).casefold())


def company_by_id(company_id: str) -> dict | None:
    with _VAULT_LOCK:
        if not vault_unlocked():
            return None
        for item in _data().get('companies', []):
            if str(item.get('id', '')) == str(company_id):
                return dict(item)
    return None


def save_company(
    company_id: str,
    name: str,
    aliases: list[str] | str,
    store: str,
    zone: str,
    brand: str,
    vendor: str,
    account_id: str,
    model_profile: str = '',
    model_prefix: str = '',
    model_min: str = '',
    model_max: str = '',
    model_pad_width: str = '',
    delivery_certificate: bool = True,
) -> str:
    name = str(name or '').strip()
    if not name:
        raise ValueError('Company name is required.')
    if isinstance(aliases, str):
        alias_list = [x.strip() for x in re.split(r'[,\n;]+', aliases) if x.strip()]
    else:
        alias_list = [str(x or '').strip() for x in aliases if str(x or '').strip()]
    store = str(store or '').strip()
    zone = str(zone or '').strip()
    brand = str(brand or '').strip().upper()
    vendor = str(vendor or '').strip().upper() or brand
    account_id = str(account_id or '').strip()
    model_profile = str(model_profile or '').strip()
    model_prefix = str(model_prefix or '').strip().rstrip('-')
    try:
        min_n = int(str(model_min).strip()) if str(model_min or '').strip() else None
        max_n = int(str(model_max).strip()) if str(model_max or '').strip() else None
        pad_n = max(0, min(8, int(str(model_pad_width).strip()))) if str(model_pad_width or '').strip() else 0
    except Exception as exc:
        raise ValueError('Model minimum, maximum and padding must be whole numbers.') from exc
    if model_prefix and (min_n is None or max_n is None):
        raise ValueError('Custom model series needs both minimum and maximum numbers.')
    if min_n is not None and max_n is not None and max_n < min_n:
        raise ValueError('Model maximum must be greater than or equal to the minimum.')
    if account_id and account_by_id(account_id) is None:
        raise ValueError('Choose a ShedSuite account that exists in the encrypted vault.')

    with _VAULT_LOCK:
        data = _data()
        companies = [x for x in data.get('companies', []) if isinstance(x, dict)]
        existing = None
        if company_id:
            existing = next((x for x in companies if str(x.get('id', '')) == str(company_id)), None)
        if existing is None:
            existing = {'id': _new_id('co')}
            companies.append(existing)
        existing.update({
            'name': name,
            'aliases': list(dict.fromkeys(alias_list)),
            'store': store,
            'zone': zone,
            'brand': brand,
            'vendor': vendor,
            'account_id': account_id,
            'model_profile': model_profile,
            'model_prefix': model_prefix,
            'model_min': min_n,
            'model_max': max_n,
            'model_pad_width': pad_n,
            'delivery_certificate': bool(delivery_certificate),
            'active': True,
        })
        data['companies'] = companies
        _save_unlocked()
        return str(existing['id'])


def delete_company(company_id: str) -> bool:
    with _VAULT_LOCK:
        data = _data()
        old = [x for x in data.get('companies', []) if isinstance(x, dict)]
        new = [x for x in old if str(x.get('id', '')) != str(company_id)]
        if len(new) == len(old):
            return False
        data['companies'] = new
        _save_unlocked()
        return True


def resolve_company_config(company: str, dealer: str = '') -> dict | None:
    """Resolve a saved company by canonical name/alias, then a conservative unique fuzzy match."""
    if not vault_unlocked():
        return None
    target = _normalize_entity(company) or _normalize_entity(dealer)
    if not target:
        return None
    candidates: list[tuple[dict, str]] = []
    for item in _data().get('companies', []):
        if not isinstance(item, dict) or item.get('active') is False:
            continue
        for alias in _company_aliases(item):
            key = _normalize_entity(alias)
            if key:
                candidates.append((item, key))
    exact = [item for item, key in candidates if key == target]
    if exact:
        return _resolved_company(exact[0])
    contained = [item for item, key in candidates if min(len(key), len(target)) >= 6 and (key in target or target in key)]
    unique_contained = {str(x.get('id')): x for x in contained}
    if len(unique_contained) == 1:
        return _resolved_company(next(iter(unique_contained.values())))
    scored: dict[str, tuple[float, dict]] = {}
    for item, key in candidates:
        score = SequenceMatcher(None, target, key).ratio()
        cid = str(item.get('id'))
        if cid not in scored or score > scored[cid][0]:
            scored[cid] = (score, item)
    ranked = sorted(scored.values(), key=lambda x: x[0], reverse=True)
    if ranked and ranked[0][0] >= 0.90 and (len(ranked) == 1 or ranked[0][0] - ranked[1][0] >= 0.04):
        return _resolved_company(ranked[0][1])
    return None


def _resolved_company(item: dict) -> dict:
    result = dict(item)
    acct = account_by_id(str(item.get('account_id', '') or ''), include_password=False)
    result['account_email'] = str((acct or {}).get('email', '') or '')
    if str(item.get('model_prefix', '') or '').strip() and item.get('model_min') is not None and item.get('model_max') is not None:
        result['dynamic_profile'] = f"VAULT_{str(item.get('id', '')).upper()}"
    else:
        result['dynamic_profile'] = ''
    return result


def credentials_for_company(company: str, dealer: str = '', explicit_email: str = '') -> tuple[str, str, str]:
    """Return (email,password,reason) deterministically. Never cycles through unrelated tenants."""
    if not vault_unlocked():
        return '', '', 'vault locked'
    config = resolve_company_config(company, dealer)
    if config and config.get('delivery_certificate') is False:
        return '', '', 'delivery certificate disabled'
    if config and config.get('account_id'):
        acct = account_by_id(str(config.get('account_id')), include_password=True)
        if acct and acct.get('email') and acct.get('password'):
            return str(acct['email']), str(acct['password']), 'company registry'
    explicit_email = str(explicit_email or '').strip()
    if explicit_email:
        acct = account_by_email(explicit_email, include_password=True)
        if acct and acct.get('password'):
            return str(acct['email']), str(acct['password']), 'existing company rule'
        return explicit_email, '', 'credential missing'
    return '', '', 'company mapping needed'


def dynamic_model_series() -> dict[str, tuple[str, int, int, bool]]:
    """Custom company series in the same tuple shape as rto_transform.MODEL_SERIES."""
    if not vault_unlocked():
        return {}
    result: dict[str, tuple[str, int, int, bool]] = {}
    for item in _data().get('companies', []):
        if not isinstance(item, dict) or item.get('active') is False:
            continue
        prefix = str(item.get('model_prefix', '') or '').strip().rstrip('-')
        lo, hi = item.get('model_min'), item.get('model_max')
        if not prefix or lo is None or hi is None:
            continue
        try:
            lo_i, hi_i = int(lo), int(hi)
        except Exception:
            continue
        pad_width = int(item.get('model_pad_width') or 0)
        key = f"VAULT_{str(item.get('id', '')).upper()}"
        result[key] = (prefix + '-', lo_i, hi_i, pad_width > 0)
    return result


def dynamic_model_padding(profile: str) -> int:
    if not vault_unlocked():
        return 0
    for item in _data().get('companies', []):
        if f"VAULT_{str(item.get('id', '')).upper()}" == str(profile):
            return int(item.get('model_pad_width') or 0)
    return 0


def import_legacy_logininfo(base: Path) -> dict:
    """One-time migration helper. Runtime certificate downloads never read XLSM."""
    from openpyxl import load_workbook

    home = Path(os.environ.get('USERPROFILE') or Path.home())
    explicit = str(os.getenv('LEGACY_XLSM_PATH', '') or '').strip()
    roots = [base, base.parent, home / 'Downloads', home / 'Desktop', home / 'Documents']
    candidates: list[Path] = [Path(explicit)] if explicit else []
    for root in roots:
        candidates.extend([
            root / 'Contract_Import.xlsm',
            root / 'Shedsuite_Import V1.1.7.6' / 'Contract_Import.xlsm',
            root / 'Shedsuite_Import V1.1.7.6' / 'Shedsuite_Import V1.1.7.6' / 'Contract_Import.xlsm',
        ])
    seen: set[str] = set()
    books: list[Path] = []
    for candidate in candidates:
        try:
            rp = candidate.expanduser().resolve()
            key = str(rp).casefold()
            if key not in seen and rp.exists() and rp.is_file():
                seen.add(key); books.append(rp)
        except Exception:
            pass
    # Bounded scan too: historical versions are often nested several folders deep.
    for root in [base.parent, base.parent.parent, home / 'Downloads', home / 'Desktop', home / 'Documents']:
        try:
            root = root.resolve()
            depth0 = len(root.parts)
            for current, dirs, files in os.walk(root):
                if len(Path(current).parts) - depth0 >= 4:
                    dirs[:] = []
                if 'Contract_Import.xlsm' in files:
                    rp = (Path(current) / 'Contract_Import.xlsm').resolve()
                    if str(rp).casefold() not in seen:
                        seen.add(str(rp).casefold()); books.append(rp)
        except Exception:
            pass
    found = 0
    imported = 0
    errors: list[str] = []
    for book in books:
        try:
            wb = load_workbook(book, data_only=True, read_only=False, keep_vba=True)
            try:
                if 'RTOPro Data' not in wb.sheetnames:
                    continue
                ws = wb['RTOPro Data']
                table = ws.tables.get('Logininfo')
                if not table:
                    continue
                rows = list(ws[table.ref])
                for row in rows[1:]:
                    email = str(row[0].value or '').strip()
                    password = str(row[1].value or '')
                    if not email or not password:
                        continue
                    found += 1
                    current = account_by_email(email)
                    if current:
                        save_account(str(current.get('id', '')), str(current.get('label', '') or ''), email, password)
                    else:
                        save_account('', email.split('@', 1)[0], email, password)
                    imported += 1
            finally:
                wb.close()
        except Exception as exc:
            errors.append(f'{book.name}: {exc}')
    return {'books': len(books), 'found': found, 'imported': imported, 'errors': errors}
